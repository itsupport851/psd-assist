from flask import Flask, request, jsonify, send_file, send_from_directory
from flask_cors import CORS
import openpyxl
import requests
import os
import io
import re
import json
import tempfile
import time
from datetime import datetime, timezone

app = Flask(__name__, static_folder='static')
CORS(app)

DROPBOX_REFRESH_TOKEN = os.environ.get('DROPBOX_REFRESH_TOKEN', '')
DROPBOX_APP_KEY = os.environ.get('DROPBOX_APP_KEY', '')
DROPBOX_APP_SECRET = os.environ.get('DROPBOX_APP_SECRET', '')
TEMPLATE_PATH = os.environ.get('TEMPLATE_PATH', '/PSD Customers/GPC_Commercial_Workbook_template.xlsm')
ANTHROPIC_API_KEY = os.environ.get('ANTHROPIC_API_KEY', '')
HUBSPOT_API_KEY = os.environ.get('HUBSPOT_API_KEY', '')
CUSTOMER_INTAKE_PIN = os.environ.get('CUSTOMER_INTAKE_PIN', '')
# Intake deals always land on the default pipeline, regardless of power company.
CUSTOMER_INTAKE_PIPELINE = 'default'
# The state on an intake submission is derived from the power company, never
# taken from the client — the form renders it read-only, but the payload is not
# trustworthy on its own.
POWER_COMPANY_STATE_MAP = {
    'Georgia Power': 'GA',
    'Entergy Louisiana': 'LA',
    'Entergy Arkansas': 'AR',
}
CUSTOMER_FORM_SENT_STAGE = os.environ.get('DEALSTAGE_CUSTOMER_FORM_SENT', 'appointmentscheduled')
UNIT_PRICE = os.environ.get('UNIT_PRICE', '0')
OWNER_ID = os.environ.get('OWNER_ID', '167151077')
CUSTOMER_INTAKE_INVITATION = 'false'
# Customer-uploaded intake documents
HUBSPOT_FILES_URL = os.environ.get('HUBSPOT_FILES_URL', 'https://api.hubapi.com/files/2026-03/files')
INTAKE_FILES_FOLDER = os.environ.get('INTAKE_FILES_FOLDER', '/customer-intake')
# HUBSPOT_DEFINED association: note -> deal. (202 is note -> contact.)
NOTE_TO_DEAL_ASSOCIATION_TYPE_ID = 214
MAX_UPLOAD_BYTES = 20 * 1024 * 1024
MAX_UPLOAD_FILES = 10
app.config['MAX_CONTENT_LENGTH'] = MAX_UPLOAD_BYTES * MAX_UPLOAD_FILES + 5 * 1024 * 1024
HUBSPOT_BASE = 'https://api.hubapi.com'
HUBSPOT_HEADERS = lambda: {'Authorization': f'Bearer {HUBSPOT_API_KEY}', 'Content-Type': 'application/json'}
PORTAL_ID = os.environ.get('PORTAL_ID', '246901747')
SERVICE_PIPELINE_ID = os.environ.get('SERVICE_PIPELINE_ID', 'ba9cdbd6-e220-45b2-a5a2-d67ebdcbade6')
SERVICE_STAGE_MAP = {
    'new': '8e2b21d0-7a90-4968-8f8c-a8525cc49c70',
    'in_progress': '600b692d-a3fe-4052-9cd7-278b134d7941',
    'closed': 'de53e7d9-6b57-4701-b576-92de01c9ed65',
}
SERVICE_PROPERTY_ALIASES = {
    'name': 'hs_name',
    'status': 'hs_status',
    'start_date': 'hs_start_date',
    'target_end_date': 'hs_target_end_date',
    'total_cost': 'hs_total_cost',
    'team_id': 'hs_shared_team_ids',
}
SERVICE_STATUS_MAP = {
    'on_track': 'on_track', 'on-track': 'on_track', 'on track': 'on_track', 'ontrack': 'on_track',
    'at_risk': 'delayed', 'at-risk': 'delayed', 'at risk': 'delayed', 'atrisk': 'delayed', 'at risk': 'delayed',
    'behind': 'failed', 'failed': 'failed',
    'complete': 'succeeded_completed', 'completed': 'succeeded_completed', 'succeeded_completed': 'succeeded_completed', 'success': 'succeeded_completed',
    'delayed': 'delayed',
}

def get_dropbox_access_token():
    response = requests.post(
        'https://api.dropbox.com/oauth2/token',
        data={
            'grant_type': 'refresh_token',
            'refresh_token': DROPBOX_REFRESH_TOKEN,
            'client_id': DROPBOX_APP_KEY,
            'client_secret': DROPBOX_APP_SECRET,
        }
    )
    if response.status_code != 200:
        raise Exception(f'Failed to refresh Dropbox token: {response.text}')
    return response.json()['access_token']

def download_template_from_dropbox():
    access_token = get_dropbox_access_token()
    url = 'https://content.dropboxapi.com/2/files/download'
    headers = {
        'Authorization': f'Bearer {access_token}',
        'Dropbox-API-Arg': f'{{"path": "{TEMPLATE_PATH}"}}'
    }
    response = requests.post(url, headers=headers)
    if response.status_code != 200:
        raise Exception(f'Failed to download template: {response.text}')
    return response.content

def fill_workbook(template_bytes, data):
    with tempfile.NamedTemporaryFile(suffix='.xlsm', delete=False) as tmp:
        tmp.write(template_bytes)
        tmp_path = tmp.name

    wb = openpyxl.load_workbook(tmp_path, keep_vba=True)
    ws = wb['Inputs']

    ws['F8'] = data.get('gp_account_name', '')
    ws['F9'] = data.get('owner_name', '')
    address_parts = [data.get('farm_address', ''), data.get('city', ''), data.get('state', ''), data.get('zip', '')]
    ws['F10'] = ', '.join(p for p in address_parts if p)
    ws['F11'] = data.get('phone', '')
    ws['F12'] = data.get('email', '')
    ws['M8'] = f"{data.get('gp_account_name', '')} — GP {data.get('gp_account_number', '')}"
    ws['F13'] = 'Agriculture'

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    os.unlink(tmp_path)
    return output

# ── PIN Authentication ──────────────────────────────────────
@app.route('/auth', methods=['POST'])
def auth():
    try:
        data = request.get_json()
        pin = str(data.get('pin', ''))
        if not pin:
            return jsonify({'error': 'PIN required'}), 400

        # Check Admin PIN
        if pin == os.environ.get('PIN_ADMIN', ''):
            return jsonify({'success': True, 'role': 'admin', 'name': 'Admin'})

        # Check Sales PIN
        if pin == os.environ.get('PIN_SALES', ''):
            return jsonify({'success': True, 'role': 'sales', 'name': 'Sales'})

        # Check Analyst PIN
        if pin == os.environ.get('PIN_ANALYST', ''):
            return jsonify({'success': True, 'role': 'analyst', 'name': 'Analyst'})

        # Check Operations PIN
        if pin == os.environ.get('PIN_OPERATIONS', ''):
            return jsonify({'success': True, 'role': 'operations', 'name': 'Operations'})

        # Check Installer PINs (PIN_INSTALLER_{team_id})
        for key, value in os.environ.items():
            if key.startswith('PIN_INSTALLER_') and pin == value:
                team_id = key.replace('PIN_INSTALLER_', '')
                try:
                    res = requests.get(f'{HUBSPOT_BASE}/settings/v3/users/teams', headers=HUBSPOT_HEADERS())
                    teams = res.json().get('results', [])
                    team_name = next((t['name'] for t in teams if str(t['id']) == team_id), 'Installer')
                except:
                    team_name = 'Installer'
                return jsonify({'success': True, 'role': 'installer', 'name': team_name, 'team_id': team_id})

        return jsonify({'error': 'Invalid PIN'}), 401
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ── Serve frontend ──────────────────────────────────────────
@app.route('/', methods=['GET'])
def index():
    return app.send_static_file('index.html')

@app.route('/sales', methods=['GET'])
def sales_portal():
    return app.send_static_file('index.html')

@app.route('/operations', methods=['GET'])
def operations_portal():
    return app.send_static_file('index.html')

@app.route('/service', methods=['GET'])
def service_portal():
    return app.send_static_file('index.html')

@app.route('/map', methods=['GET'])
def map_page():
    return app.send_static_file('map.html')

@app.route('/customer-intake', methods=['GET'])
def customer_intake_page():
    return app.send_static_file('customer-intake.html')

@app.route('/Images/<path:filename>', methods=['GET'])
def image_asset(filename):
    filename = filename.replace('\\', '/')
    for directory in ['Images', os.path.join(app.static_folder, 'Images')]:
        candidate = os.path.abspath(os.path.join(directory, filename))
        if candidate.startswith(os.path.abspath(directory) + os.sep) and os.path.isfile(candidate):
            return send_from_directory(directory, filename)
    return jsonify({'error': 'Image asset not found'}), 404

@app.route('/Docs/<path:filename>', methods=['GET'])
def document_asset(filename):
    filename = filename.replace('\\', '/')
    for directory in ['Docs', os.path.join(app.static_folder, 'Docs')]:
        candidate = os.path.abspath(os.path.join(directory, filename))
        if candidate.startswith(os.path.abspath(directory) + os.sep) and os.path.isfile(candidate):
            return send_from_directory(directory, filename)
    return jsonify({'error': 'Document asset not found'}), 404

@app.route('/customer-intake-auth', methods=['POST'])
def customer_intake_auth():
    try:
        data = request.get_json()
        pin = str(data.get('pin', ''))
        if not pin:
            return jsonify({'error': 'PIN required'}), 400
        if not CUSTOMER_INTAKE_PIN:
            return jsonify({'error': 'Form is not yet configured'}), 500
        if pin != CUSTOMER_INTAKE_PIN:
            return jsonify({'error': 'Invalid PIN'}), 401
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/maps-key', methods=['GET'])
def maps_key():
    return jsonify({'key': os.environ.get('GOOGLE_MAPS_API_KEY', '')})

# ── Health check ────────────────────────────────────────────
@app.route('/health', methods=['GET'])
def health():
    return jsonify({'status': 'ok'})

# ── Fill workbook ────────────────────────────────────────────
@app.route('/fill-workbook', methods=['POST'])
def fill_workbook_endpoint():
    try:
        data = request.get_json()
        if not data:
            return jsonify({'error': 'No JSON data provided'}), 400
        required = ['gp_account_name', 'gp_account_number']
        missing = [f for f in required if not data.get(f)]
        if missing:
            return jsonify({'error': f'Missing required fields: {missing}'}), 400
        template_bytes = download_template_from_dropbox()
        filled_workbook = fill_workbook(template_bytes, data)
        safe_name = data.get('gp_account_name', 'Customer').replace('/', '-').replace('\\', '-')
        filename = f"GPC_Commercial_Workbook_{safe_name}.xlsm"
        return send_file(filled_workbook, mimetype='application/vnd.ms-excel.sheet.macroEnabled.12', as_attachment=True, download_name=filename)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ── HubSpot: Get open deals ──────────────────────────────────
@app.route('/hubspot/deals', methods=['GET'])
def hs_get_deals():
    try:
        props = 'dealname,amount,dealstage,closedate,hs_object_id,total_fans,total_barns,operation_type'
        stage_filter = request.args.get('stage', None)
        url = f'{HUBSPOT_BASE}/crm/v3/objects/deals?limit=50&properties={props}'
        res = requests.get(url, headers=HUBSPOT_HEADERS())
        if res.status_code != 200:
            return jsonify({'error': res.text}), res.status_code
        data = res.json()
        deals = []
        for d in data.get('results', []):
            props_data = d.get('properties', {})
            if stage_filter and props_data.get('dealstage') != stage_filter:
                continue
            deals.append({
                'id': d['id'],
                'properties': props_data
            })
        return jsonify({'deals': deals})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ── HubSpot: Get line items for a deal ──────────────────────
@app.route('/hubspot/line-items/<deal_id>', methods=['GET'])
def hs_get_line_items(deal_id):
    try:
        # Get associated line item IDs
        assoc_url = f'{HUBSPOT_BASE}/crm/v3/objects/deals/{deal_id}/associations/line_items'
        assoc_res = requests.get(assoc_url, headers=HUBSPOT_HEADERS())
        if assoc_res.status_code != 200:
            return jsonify({'error': assoc_res.text}), assoc_res.status_code

        assoc_data = assoc_res.json()
        line_item_ids = [r['id'] for r in assoc_data.get('results', [])]

        if not line_item_ids:
            return jsonify({'line_items': []})

        # Fetch each line item
        props = 'name,price,quantity,description,fan_brand,fan_size,motor_size,number_of_fans,hs_object_id'
        line_items = []
        for lid in line_item_ids:
            li_res = requests.get(f'{HUBSPOT_BASE}/crm/v3/objects/line_items/{lid}?properties={props}', headers=HUBSPOT_HEADERS())
            if li_res.status_code == 200:
                li = li_res.json()
                line_items.append({
                    'id': li['id'],
                    'properties': li.get('properties', {})
                })

        return jsonify({'line_items': line_items})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ── HubSpot: Get deals with addresses ───────────────────────
@app.route('/hubspot/deals-with-addresses', methods=['GET'])
def hs_deals_with_addresses():
    try:
        stage_filter = request.args.get('stage', None)  # None = all deals
        props = 'dealname,amount,dealstage,total_fans,total_barns,hs_object_id'
        url = f'{HUBSPOT_BASE}/crm/v3/objects/deals?limit=100&properties={props}'
        res = requests.get(url, headers=HUBSPOT_HEADERS())
        if res.status_code != 200:
            return jsonify({'error': res.text}), res.status_code

        all_deals = res.json().get('results', [])
        if stage_filter:
            all_deals = [d for d in all_deals if d.get('properties', {}).get('dealstage') == stage_filter]
        else:
            # Exclude closed lost by default; include everything else
            all_deals = [d for d in all_deals if d.get('properties', {}).get('dealstage') != 'closedlost']

        result = []
        contact_props = 'address,city,state,zip,company,firstname,ower_name,email,phone,georgia_power_account'

        for deal in all_deals:
            deal_id = deal['id']
            assoc_res = requests.get(f'{HUBSPOT_BASE}/crm/v3/objects/deals/{deal_id}/associations/contacts', headers=HUBSPOT_HEADERS())
            if assoc_res.status_code != 200:
                continue
            contacts = assoc_res.json().get('results', [])
            if not contacts:
                continue

            contact_id = contacts[0]['id']
            contact_res = requests.get(f'{HUBSPOT_BASE}/crm/v3/objects/contacts/{contact_id}?properties={contact_props}', headers=HUBSPOT_HEADERS())
            if contact_res.status_code != 200:
                continue

            contact = contact_res.json().get('properties', {})
            dp = deal.get('properties', {})

            result.append({
                'deal_id': deal_id,
                'deal_name': dp.get('dealname', ''),
                'dealstage': dp.get('dealstage', ''),
                'amount': dp.get('amount', ''),
                'total_fans': dp.get('total_fans', ''),
                'total_barns': dp.get('total_barns', ''),
                'company': contact.get('company', ''),
                'owner': contact.get('ower_name', '') or contact.get('firstname', ''),
                'address': contact.get('address', ''),
                'city': contact.get('city', ''),
                'state': contact.get('state', ''),
                'zip': contact.get('zip', ''),
                'email': contact.get('email', ''),
                'phone': contact.get('phone', '')
            })

        return jsonify({'deals_with_addresses': result})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ── HubSpot: Update line item ───────────────────────────────
@app.route('/hubspot/line-items/<line_item_id>', methods=['PATCH'])
def hs_update_line_item(line_item_id):
    try:
        data = request.get_json()
        payload = {'properties': {}}
        if 'unit_price' in data:
            payload['properties']['price'] = str(data['unit_price'])
        if 'quantity' in data:
            payload['properties']['quantity'] = str(data['quantity'])
        if 'name' in data:
            payload['properties']['name'] = data['name']
        res = requests.patch(f'{HUBSPOT_BASE}/crm/v3/objects/line_items/{line_item_id}', json=payload, headers=HUBSPOT_HEADERS())
        if res.status_code not in [200, 204]:
            return jsonify({'error': res.text}), res.status_code
        return jsonify({'success': True, 'line_item_id': line_item_id})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ── HubSpot: Update deal ─────────────────────────────────────
@app.route('/hubspot/deals/<deal_id>', methods=['PATCH'])
def hs_update_deal(deal_id):
    try:
        data = request.get_json()
        payload = {'properties': {}}
        if 'amount' in data:
            payload['properties']['amount'] = str(data['amount'])
        if 'dealstage' in data:
            payload['properties']['dealstage'] = data['dealstage']
        if 'dealname' in data:
            payload['properties']['dealname'] = data['dealname']
        res = requests.patch(f'{HUBSPOT_BASE}/crm/v3/objects/deals/{deal_id}', json=payload, headers=HUBSPOT_HEADERS())
        if res.status_code not in [200, 204]:
            return jsonify({'error': res.text}), res.status_code
        return jsonify({'success': True, 'deal_id': deal_id})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ── HubSpot: Add line item to deal ──────────────────────────
@app.route('/hubspot/line-items', methods=['POST'])
def hs_create_line_item():
    try:
        data = request.get_json()
        deal_id = data.get('deal_id')
        if not deal_id:
            return jsonify({'error': 'deal_id is required'}), 400
        payload = {
            'properties': {
                'name': data.get('name', 'Fan Motor'),
                'price': str(data.get('unit_price', 0)),
                'quantity': str(data.get('quantity', 1)),
                'description': data.get('description', ''),
                'fan_brand': data.get('fan_brand', ''),
                'fan_size': str(data.get('fan_size', '')),
                'motor_size': str(data.get('motor_size', '')),
            }
        }
        res = requests.post(f'{HUBSPOT_BASE}/crm/v3/objects/line_items', json=payload, headers=HUBSPOT_HEADERS())
        if res.status_code != 201:
            return jsonify({'error': res.text}), res.status_code
        line_item_id = res.json()['id']
        assoc_payload = {'inputs': [{'from': {'id': deal_id}, 'to': {'id': line_item_id}, 'type': 'deal_to_line_item'}]}
        requests.post(f'{HUBSPOT_BASE}/crm/v3/associations/deals/line_items/batch/create', json=assoc_payload, headers=HUBSPOT_HEADERS())
        return jsonify({'success': True, 'line_item_id': line_item_id})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ── HubSpot: Create new deal ─────────────────────────────────
@app.route('/hubspot/deals', methods=['POST'])
def hs_create_deal():
    try:
        data = request.get_json()
        payload = {
            'properties': {
                'dealname': data.get('dealname', ''),
                'dealstage': data.get('dealstage', 'appointmentscheduled'),
                'pipeline': 'default',
                'amount': str(data.get('amount', 0)),
                'psd_gp_account_name': data.get('gp_account_name', ''),
                'psd_gp_account_number': data.get('gp_account_number', ''),
                'psd_owner_name': data.get('owner_name', ''),
                'psd_farm_address': data.get('farm_address', ''),
                'psd_city': data.get('city', ''),
                'psd_state': data.get('state', ''),
                'psd_zip': data.get('zip', ''),
                'total_fans': str(data.get('total_fans', '')),
                'total_barns': str(data.get('total_barns', '')),
            }
        }
        res = requests.post(f'{HUBSPOT_BASE}/crm/v3/objects/deals', json=payload, headers=HUBSPOT_HEADERS())
        if res.status_code != 201:
            return jsonify({'error': res.text}), res.status_code
        deal = res.json()
        return jsonify({'success': True, 'deal_id': deal['id'], 'dealname': deal['properties'].get('dealname')})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ── HubSpot: Create new contact ──────────────────────────────
@app.route('/hubspot/contacts', methods=['POST'])
def hs_create_contact():
    try:
        data = request.get_json()
        email = str(data.get('email', '')).strip()
        if not email:
            return jsonify({'error': 'email is required'}), 400
        invitation_input = str(data.get('invitation', '')).strip().lower()
        if invitation_input in ['yes', 'true', '1']:
            invitation = 'true'
        elif invitation_input in ['no', 'false', '0']:
            invitation = 'false'
        else:
            return jsonify({'error': 'invitation must be true or false'}), 400

        payload = {
            'properties': {
                'email': email,
                'firstname': data.get('first_name', data.get('firstname', '')),
                'lastname': data.get('last_name', data.get('lastname', '')),
                'state': data.get('state', ''),
                'invitation': invitation,
            }
        }

        search_payload = {
            'filterGroups': [{'filters': [{'propertyName': 'email', 'operator': 'EQ', 'value': email}]}],
            'properties': ['email', 'firstname', 'lastname', 'state', 'invitation'],
            'limit': 1
        }
        search_res = requests.post(f'{HUBSPOT_BASE}/crm/v3/objects/contacts/search', json=search_payload, headers=HUBSPOT_HEADERS())
        existing = search_res.json().get('results', []) if search_res.status_code == 200 else []

        if existing:
            contact_id = existing[0]['id']
            res = requests.patch(f'{HUBSPOT_BASE}/crm/v3/objects/contacts/{contact_id}', json=payload, headers=HUBSPOT_HEADERS())
            if res.status_code not in [200, 204]:
                return jsonify({'error': res.text}), res.status_code
            contact = {'id': contact_id, 'properties': payload['properties']}
            updated = True
        else:
            res = requests.post(f'{HUBSPOT_BASE}/crm/v3/objects/contacts', json=payload, headers=HUBSPOT_HEADERS())
            if res.status_code != 201:
                return jsonify({'error': res.text}), res.status_code
            contact = res.json()
            updated = False

        deal_id = data.get('deal_id')
        if deal_id:
            assoc_payload = {'inputs': [{'from': {'id': deal_id}, 'to': {'id': contact['id']}, 'type': 'deal_to_contact'}]}
            requests.post(f'{HUBSPOT_BASE}/crm/v3/associations/deals/contacts/batch/create', json=assoc_payload, headers=HUBSPOT_HEADERS())
        return jsonify({'success': True, 'updated': updated, 'contact_id': contact['id'], 'email': email})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ── HubSpot: Get teams ───────────────────────────────────────
@app.route('/hubspot/teams', methods=['GET'])
def hs_get_teams():
    try:
        res = requests.get(f'{HUBSPOT_BASE}/settings/v3/users/teams', headers=HUBSPOT_HEADERS())
        if res.status_code != 200:
            return jsonify({'error': res.text}), res.status_code
        teams = []
        for t in res.json().get('results', []):
            teams.append({
                'id': t.get('id'),
                'name': t.get('name'),
                'userIds': t.get('userIds', [])
            })
        return jsonify({'teams': teams})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ── HubSpot: Get service pipeline stages ────────────────────
def resolve_service_stage(stage_value):
    if not stage_value:
        return SERVICE_STAGE_MAP['new']
    if stage_value in SERVICE_STAGE_MAP:
        return SERVICE_STAGE_MAP[stage_value]
    if stage_value in SERVICE_STAGE_MAP.values():
        return stage_value
    return SERVICE_STAGE_MAP['new']


def normalize_service_name(name):
    if not name:
        return name
    name = re.sub(r'\b(?:GP\s+)+GP\b', 'GP', name)
    name = re.sub(r'\s+', ' ', name).strip()
    return name


def build_service_properties(data, include_pipeline=False):
    properties = {}
    for key, hs_name in SERVICE_PROPERTY_ALIASES.items():
        if key in data and data.get(key) is not None and data.get(key) != "":
            if key == 'name':
                properties[hs_name] = normalize_service_name(str(data[key]))
            elif key == 'status':
                raw_status = str(data[key]).strip()
                normalized_status = raw_status.lower().replace('-', '_').replace(' ', '_')
                properties[hs_name] = SERVICE_STATUS_MAP.get(normalized_status, SERVICE_STATUS_MAP.get(raw_status, raw_status))
            else:
                properties[hs_name] = str(data[key])
    if include_pipeline:
        properties['hs_pipeline'] = SERVICE_PIPELINE_ID
        properties['hs_pipeline_stage'] = resolve_service_stage(data.get('stage', 'new'))
    return properties


def get_deal_contact_ids(deal_id):
    try:
        res = requests.get(f'{HUBSPOT_BASE}/crm/v3/objects/deals/{deal_id}/associations/contacts', headers=HUBSPOT_HEADERS())
        if res.status_code == 200:
            return [r['id'] for r in res.json().get('results', []) if r.get('id')]
    except:
        pass
    return []


def associate_service_contacts(service_id, contact_ids):
    if not contact_ids:
        return
    payload = {
        'inputs': [
            {'from': {'id': service_id}, 'to': {'id': cid}, 'type': 'service_to_contact'}
            for cid in contact_ids
        ]
    }
    try:
        requests.post(f'{HUBSPOT_BASE}/crm/v3/associations/services/contacts/batch/create', json=payload, headers=HUBSPOT_HEADERS())
    except:
        pass


def create_service_task(service_id, task_details, contact_ids=None):
    if not task_details:
        return None
    payload = {
        'properties': {
            'subject': 'Service Task',
            'hs_task_body': str(task_details),
        }
    }
    try:
        res = requests.post(f'{HUBSPOT_BASE}/crm/v3/objects/tasks', json=payload, headers=HUBSPOT_HEADERS())
        if res.status_code != 201:
            return None
        task_id = res.json().get('id')
        if task_id:
            # Attach the created task to the service. Try both directions to handle HubSpot association expectations.
            assoc_inputs = [
                {'from': {'id': service_id}, 'to': {'id': task_id}, 'type': 'service_to_task'}
            ]
            assoc_res = requests.post(
                f'{HUBSPOT_BASE}/crm/v3/associations/services/tasks/batch/create',
                json={'inputs': assoc_inputs},
                headers=HUBSPOT_HEADERS()
            )
            if assoc_res.status_code not in [200, 201]:
                requests.post(
                    f'{HUBSPOT_BASE}/crm/v3/associations/tasks/services/batch/create',
                    json={'inputs': [{'from': {'id': task_id}, 'to': {'id': service_id}, 'type': 'task_to_service'}]},
                    headers=HUBSPOT_HEADERS()
                )
            if contact_ids:
                requests.post(
                    f'{HUBSPOT_BASE}/crm/v3/associations/tasks/contacts/batch/create',
                    json={'inputs': [{'from': {'id': task_id}, 'to': {'id': cid}, 'type': 'task_to_contact'} for cid in contact_ids]},
                    headers=HUBSPOT_HEADERS()
                )
        return task_id
    except:
        return None


def _get_results_ids(path):
    try:
        res = requests.get(path, headers=HUBSPOT_HEADERS())
        if res.status_code != 200:
            return []
        return [r.get('id') for r in res.json().get('results', []) if r.get('id')]
    except:
        return []


def _parse_team_ids(team_value):
    if not team_value:
        return []
    if isinstance(team_value, list):
        return [str(x).strip() for x in team_value if x]
    return [x.strip() for x in str(team_value).split(';') if x.strip()]


def _service_name_from_props(props):
    raw_name = (
        props.get('hs_name') or
        props.get('hs_object_name') or
        props.get('subject') or
        props.get('hs_ticket_name') or
        props.get('content') or
        props.get('title') or
        props.get('name') or
        ''
    ).strip()
    if not raw_name:
        raw_name = (props.get('hs_ticket_category') or '').strip()
    return raw_name


def _service_stage_label(stage_id):
    if not stage_id:
        return 'New'
    stage_map = {
        SERVICE_STAGE_MAP['new']: 'New',
        SERVICE_STAGE_MAP['in_progress']: 'In Progress',
        SERVICE_STAGE_MAP['closed']: 'Closed'
    }
    return stage_map.get(stage_id, stage_id)


def _service_status_label(status):
    if not status:
        return ''
    status_map = {
        'ON_TRACK': 'On Track', 'on_track': 'On Track', 'On Track': 'On Track',
        'AT_RISK': 'At Risk', 'at_risk': 'At Risk', 'At Risk': 'At Risk',
        'BEHIND': 'Behind', 'Behind': 'Behind', 'COMPLETE': 'Completed', 'Completed': 'Completed',
        'delayed': 'At Risk', 'failed': 'Behind', 'succeeded_completed': 'Completed'
    }
    return status_map.get(status, status)


def _service_details_from_props(service_id, props):
    team_id = props.get('hs_shared_team_ids') or props.get('hubspot_team_id') or ''
    return {
        'id': service_id,
        'name': _service_name_from_props(props),
        'description': props.get('hs_description', ''),
        'deal_name': '',
        'stage': _service_stage_label(props.get('hs_pipeline_stage', '')),
        'stage_id': props.get('hs_pipeline_stage', ''),
        'status': _service_status_label(props.get('hs_status') or props.get('hs_ticket_priority') or props.get('hs_object_status') or ''),
        'team': team_id,
        'team_id': team_id,
        'start_date': props.get('hs_start_date', ''),
        'due_date': props.get('hs_target_end_date', ''),
        'total_cost': props.get('hs_total_cost', ''),
        'amount_paid': props.get('hs_amount_paid', ''),
        'remaining': props.get('hs_remaining_amount', '')
    }


@app.route('/hubspot/team-services/<team_id>', methods=['GET'])
def hs_team_services(team_id):
    try:
        props = 'hs_name,hs_description,hs_pipeline,hs_pipeline_stage,hs_status,hs_start_date,hs_target_end_date,hs_total_cost,hs_shared_team_ids,hubspot_team_id,hs_object_name,subject,hs_ticket_name,hs_ticket_category'
        url = f'{HUBSPOT_BASE}/crm/v3/objects/services?limit=100&properties={props}'
        res = requests.get(url, headers=HUBSPOT_HEADERS())
        if res.status_code != 200:
            return jsonify({'error': res.text}), res.status_code

        services = []
        for raw in res.json().get('results', []):
            props = raw.get('properties', {})
            team_ids = _parse_team_ids(props.get('hs_shared_team_ids') or props.get('hubspot_team_id'))
            if str(team_id) in team_ids:
                services.append(_service_details_from_props(raw.get('id'), props))

        return jsonify({'services': services})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/hubspot/service-details/<service_id>', methods=['GET'])
def hs_service_details(service_id):
    try:
        props = 'hs_name,hs_description,hs_pipeline,hs_pipeline_stage,hs_status,hs_start_date,hs_target_end_date,hs_total_cost,hs_amount_paid,hs_remaining_amount,hs_shared_team_ids,hubspot_team_id,hs_object_name,subject,hs_ticket_name,hs_ticket_category,createdate'
        url = f'{HUBSPOT_BASE}/crm/v3/objects/services/{service_id}?properties={props}'
        res = requests.get(url, headers=HUBSPOT_HEADERS())
        if res.status_code != 200:
            return jsonify({'error': res.text}), res.status_code

        raw = res.json()
        svc = _service_details_from_props(service_id, raw.get('properties', {}))

        contacts = []
        for contact_id in _get_results_ids(f'{HUBSPOT_BASE}/crm/v3/objects/services/{service_id}/associations/contacts'):
            contact_res = requests.get(f'{HUBSPOT_BASE}/crm/v3/objects/contacts/{contact_id}?properties=firstname,lastname,company,email,phone', headers=HUBSPOT_HEADERS())
            if contact_res.status_code == 200:
                cp = contact_res.json().get('properties', {})
                contacts.append({
                    'id': contact_id,
                    'name': ((cp.get('firstname') or '') + ' ' + (cp.get('lastname') or '')).strip() or cp.get('company') or 'Contact',
                    'company': cp.get('company', ''),
                    'email': cp.get('email', ''),
                    'phone': cp.get('phone', '')
                })

        tasks = []
        for task_id in _get_results_ids(f'{HUBSPOT_BASE}/crm/v3/objects/services/{service_id}/associations/tasks'):
            task_res = requests.get(
                f'{HUBSPOT_BASE}/crm/v3/objects/tasks/{task_id}?properties=subject,hs_task_body,hs_task_status,hs_task_priority,hs_task_start_date,hs_task_due_date,createdate',
                headers=HUBSPOT_HEADERS()
            )
            if task_res.status_code == 200:
                tp = task_res.json().get('properties', {})
                tasks.append({
                    'id': task_id,
                    'subject': tp.get('subject', ''),
                    'body': tp.get('hs_task_body', ''),
                    'status': tp.get('hs_task_status', ''),
                    'priority': tp.get('hs_task_priority', ''),
                    'start_date': tp.get('hs_task_start_date', tp.get('createdate', '')),
                    'due_date': tp.get('hs_task_due_date', ''),
                    'created_date': tp.get('createdate', '')
                })

        deals = []
        for deal_id in _get_results_ids(f'{HUBSPOT_BASE}/crm/v3/objects/services/{service_id}/associations/deals'):
            deal_res = requests.get(f'{HUBSPOT_BASE}/crm/v3/objects/deals/{deal_id}?properties=dealname,amount,dealstage', headers=HUBSPOT_HEADERS())
            if deal_res.status_code == 200:
                dp = deal_res.json().get('properties', {})
                deals.append({
                    'id': deal_id,
                    'name': dp.get('dealname', ''),
                    'amount': dp.get('amount', ''),
                    'stage': dp.get('dealstage', '')
                })

        return jsonify({'service': svc, 'contacts': contacts, 'tasks': tasks, 'deals': deals})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/hubspot/service-stages', methods=['GET'])
def hs_service_stages():
    try:
        res = requests.get(f'{HUBSPOT_BASE}/crm/v3/pipelines/services/{SERVICE_PIPELINE_ID}/stages', headers=HUBSPOT_HEADERS())
        if res.status_code != 200:
            return jsonify({'error': res.text}), res.status_code
        stages = [{'id': s['id'], 'label': s['label']} for s in res.json().get('results', [])]
        return jsonify({'stages': stages})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ── HubSpot: Get full service record (debug) ────────────────
@app.route('/hubspot/service-debug/<service_id>', methods=['GET'])
def hs_service_debug(service_id):
    try:
        # First get all available properties for services
        props_res = requests.get(f'{HUBSPOT_BASE}/crm/v3/properties/services', headers=HUBSPOT_HEADERS())
        all_props = [p['name'] for p in props_res.json().get('results', [])]
        prop_str = ','.join(all_props[:50])  # fetch first 50 props

        # Fetch the service with all properties
        res = requests.get(f'{HUBSPOT_BASE}/crm/v3/objects/services/{service_id}?properties={prop_str}', headers=HUBSPOT_HEADERS())
        if res.status_code != 200:
            return jsonify({'error': res.text}), res.status_code

        # Return only non-null properties
        props = res.json().get('properties', {})
        non_null = {k: v for k, v in props.items() if v is not None and v != ''}
        return jsonify({'id': service_id, 'properties': non_null, 'all_available_props': all_props})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ── HubSpot: Get service properties (debug) ─────────────────
@app.route('/hubspot/service-properties', methods=['GET'])
def hs_service_properties():
    try:
        res = requests.get(f'{HUBSPOT_BASE}/crm/v3/properties/services', headers=HUBSPOT_HEADERS())
        if res.status_code != 200:
            return jsonify({'error': res.text}), res.status_code
        props = [{'name': p['name'], 'label': p['label'], 'type': p['type']} for p in res.json().get('results', [])]
        return jsonify({'properties': props})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ── HubSpot: Get ALL services across all deals ───────────────
@app.route('/hubspot/all-services', methods=['GET'])
def hs_get_all_services():
    try:
        # 1. Get all deals (we need deal names to enrich services)
        deal_props = 'dealname'
        deals_res = requests.get(f'{HUBSPOT_BASE}/crm/v3/objects/deals?limit=100&properties={deal_props}', headers=HUBSPOT_HEADERS())
        deal_map = {}
        if deals_res.status_code == 200:
            for d in deals_res.json().get('results', []):
                deal_map[d['id']] = d.get('properties', {}).get('dealname', '')

        # 2. Fetch pipeline stage labels so we can map IDs → names
        pipeline_id = SERVICE_PIPELINE_ID
        stage_map = {}  # id -> label
        try:
            stages_res = requests.get(f'{HUBSPOT_BASE}/crm/v3/pipelines/services/{pipeline_id}/stages', headers=HUBSPOT_HEADERS())
            if stages_res.status_code == 200:
                for st in stages_res.json().get('results', []):
                    stage_map[st['id']] = st['label']
        except:
            pass
        # Fallback common internal IDs used by HubSpot default service pipeline
        stage_map.setdefault('new',         'New')
        stage_map.setdefault('in_progress', 'In Progress')
        stage_map.setdefault('closed',      'Closed')
        stage_map.setdefault('1',           'New')
        stage_map.setdefault('2',           'In Progress')
        stage_map.setdefault('3',           'Closed')

        # 3. Get all services — fetch broad property set including hs_object_name
        svc_props = 'hs_name,hs_description,hs_pipeline,hs_pipeline_stage,hs_status,hs_target_end_date,hs_start_date,hs_total_cost,hs_amount_paid,hs_amount_remaining,hs_shared_team_ids,hubspot_team_id,hs_object_name,subject,hs_ticket_name,hs_object_status,hs_ticket_priority,hs_ticket_category,createdate'
        svc_res = requests.get(f'{HUBSPOT_BASE}/crm/v3/objects/services?limit=100&properties={svc_props}', headers=HUBSPOT_HEADERS())
        if svc_res.status_code != 200:
            return jsonify({'error': svc_res.text}), svc_res.status_code

        services_raw = svc_res.json().get('results', [])

        # 4. Batch-fetch all deal associations in ONE call (avoids per-service rate limits)
        service_ids = [s['id'] for s in services_raw]
        service_to_deal = {}
        try:
            batch_payload = {'inputs': [{'id': sid} for sid in service_ids]}
            batch_res = requests.post(
                f'{HUBSPOT_BASE}/crm/v3/associations/services/deals/batch/read',
                json=batch_payload, headers=HUBSPOT_HEADERS()
            )
            if batch_res.status_code == 200:
                for result in batch_res.json().get('results', []):
                    from_id = str(result.get('from', {}).get('id', ''))
                    to_ids = [str(r['id']) for r in result.get('to', [])]
                    if from_id and to_ids:
                        service_to_deal[from_id] = to_ids[0]
        except:
            pass

        # 5. Build service records
        services = []
        for s in services_raw:
            sid = s['id']
            sp = s.get('properties', {})
            deal_name = deal_map.get(service_to_deal.get(sid, ''), '')

            # Team: HubSpot Services stores team in hs_shared_team_ids (shared teams)
            # also check hubspot_team_id (assigned team) as fallback
            team_id = sp.get('hs_shared_team_ids', '') or sp.get('hubspot_team_id', '') or ''
            team_name = team_id  # enriched below from teams API

            # Resolve name — try every known field HubSpot might store service name in
            raw_name = (
                sp.get('hs_name') or
                sp.get('hs_object_name') or
                sp.get('subject') or
                sp.get('hs_ticket_name') or
                sp.get('content') or
                sp.get('title') or
                sp.get('name') or
                ''
            ).strip()
            if not raw_name:
                raw_name = (sp.get('hs_ticket_category') or '').strip()

            # Resolve pipeline stage ID → label
            raw_stage = sp.get('hs_pipeline_stage', '') or ''
            stage_label = stage_map.get(raw_stage, raw_stage)

            # Status: use the service-specific status field first, with fallbacks
            raw_status = (
                sp.get('hs_status') or
                sp.get('hs_ticket_priority') or
                sp.get('hs_object_status') or
                ''
            )
            # Normalize HubSpot values into the app's internal status labels
            status_map = {
                'on_track': 'ON_TRACK', 'ON TRACK': 'ON_TRACK', 'On Track': 'ON_TRACK', 'on track': 'ON_TRACK', 'ON_TRACK': 'ON_TRACK',
                'delayed': 'AT_RISK', 'At Risk': 'AT_RISK', 'at_risk': 'AT_RISK', 'AT_RISK': 'AT_RISK', 'at risk': 'AT_RISK',
                'failed': 'BEHIND', 'Behind': 'BEHIND', 'BEHIND': 'BEHIND',
                'succeeded_completed': 'COMPLETE', 'Completed': 'COMPLETE', 'COMPLETE': 'COMPLETE',
            }
            raw_status = status_map.get(raw_status, raw_status)

            def normalize_date(val):
                """Convert HubSpot date value to ISO date string regardless of format."""
                if not val:
                    return ''
                try:
                    ts = float(val)
                    from datetime import datetime, timezone
                    return datetime.fromtimestamp(ts / 1000, tz=timezone.utc).strftime('%Y-%m-%d')
                except (ValueError, TypeError):
                    return str(val)[:10]

            start_raw = sp.get('start_date') or sp.get('hs_start_date') or sp.get('createdate') or ''
            # Due date: "Target End Date" field in HubSpot UI = target_end_date property
            due_raw = sp.get('target_end_date') or sp.get('hs_due_date') or ''

            services.append({
                'id': sid,
                'name': raw_name,
                'description': sp.get('description', ''),
                'deal_name': deal_name,
                'stage': stage_label,
                'stage_id': raw_stage,
                'status': raw_status,
                'team': team_name,
                'team_id': team_id,
                'start_date': normalize_date(start_raw),
                'due_date':   normalize_date(due_raw),
                'total_cost': sp.get('hs_total_cost', ''),
                'amount_paid': sp.get('hs_amount_paid', ''),
                'remaining': sp.get('hs_remaining_amount', ''),
            })

        # 5. Enrich team names in one call
        try:
            teams_res = requests.get(f'{HUBSPOT_BASE}/settings/v3/users/teams', headers=HUBSPOT_HEADERS())
            if teams_res.status_code == 200:
                team_map = {str(t['id']): t['name'] for t in teams_res.json().get('results', [])}
                for svc in services:
                    raw_tid = svc['team_id']
                    if raw_tid:
                        # hs_shared_team_ids may be semicolon-separated list of IDs
                        ids = [x.strip() for x in str(raw_tid).split(';') if x.strip()]
                        names = [team_map.get(i, i) for i in ids]
                        svc['team'] = ', '.join(names) if names else raw_tid
        except:
            pass

        return jsonify({'services': services, 'count': len(services)})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ── HubSpot: Repair service names (backfill from category/description) ──
@app.route('/hubspot/services/<service_id>/repair', methods=['PATCH'])
def hs_repair_service(service_id):
    try:
        data = request.get_json()
        name = data.get('name', '')
        if not name:
            return jsonify({'error': 'name is required'}), 400
        payload = {
            'properties': {
                'hs_name':        name,
                'hs_object_name': name,
                'subject':        name,
                'hs_ticket_name': name,
                'content':        name,
            }
        }
        res = requests.patch(f'{HUBSPOT_BASE}/crm/v3/objects/services/{service_id}', json=payload, headers=HUBSPOT_HEADERS())
        if res.status_code not in [200, 204]:
            return jsonify({'error': res.text}), res.status_code
        return jsonify({'success': True, 'service_id': service_id, 'name': name})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ── HubSpot: Inspect a service — fetch every non-null property ──
@app.route('/hubspot/service-inspect/<service_id>', methods=['GET'])
def hs_service_inspect(service_id):
    try:
        # Get all available property names for services
        props_res = requests.get(f'{HUBSPOT_BASE}/crm/v3/properties/services', headers=HUBSPOT_HEADERS())
        all_prop_names = [p['name'] for p in props_res.json().get('results', [])]

        # Fetch the service in batches of 50 props
        all_values = {}
        for i in range(0, len(all_prop_names), 50):
            batch = all_prop_names[i:i+50]
            prop_str = ','.join(batch)
            res = requests.get(f'{HUBSPOT_BASE}/crm/v3/objects/services/{service_id}?properties={prop_str}', headers=HUBSPOT_HEADERS())
            if res.status_code == 200:
                for k, v in res.json().get('properties', {}).items():
                    if v is not None and v != '':
                        all_values[k] = v

        return jsonify({'service_id': service_id, 'properties': all_values})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ── HubSpot: Get services for a deal ────────────────────────
@app.route('/hubspot/services/<deal_id>', methods=['GET'])
def hs_get_services(deal_id):
    try:
        # Get associated service IDs
        assoc_res = requests.get(f'{HUBSPOT_BASE}/crm/v3/objects/deals/{deal_id}/associations/services', headers=HUBSPOT_HEADERS())
        if assoc_res.status_code != 200:
            return jsonify({'services': []})
        service_ids = [r['id'] for r in assoc_res.json().get('results', [])]
        if not service_ids:
            return jsonify({'services': []})
        props = 'hs_name,hs_description,hs_pipeline,hs_pipeline_stage,hs_status,hs_start_date,hs_target_end_date,hs_total_cost,hs_shared_team_ids'
        services = []
        for sid in service_ids:
            res = requests.get(f'{HUBSPOT_BASE}/crm/v3/objects/services/{sid}?properties={props}', headers=HUBSPOT_HEADERS())
            if res.status_code == 200:
                s = res.json()
                services.append({'id': s['id'], 'properties': s.get('properties', {})})
        return jsonify({'services': services})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ── HubSpot: Create service for a deal ──────────────────────
@app.route('/hubspot/services', methods=['POST'])
def hs_create_service():
    try:
        data = request.get_json()
        deal_id = data.get('deal_id')
        if not deal_id:
            return jsonify({'error': 'deal_id is required'}), 400

        task_details = data.get('task_details') or data.get('description', '')
        payload = {'properties': build_service_properties(data, include_pipeline=True)}

        res = requests.post(f'{HUBSPOT_BASE}/crm/v3/objects/services', json=payload, headers=HUBSPOT_HEADERS())
        if res.status_code != 201:
            return jsonify({'error': res.text}), res.status_code

        service_id = res.json()['id']

        # Associate service to deal
        assoc_payload = {'inputs': [{'from': {'id': deal_id}, 'to': {'id': service_id}, 'type': 'deal_to_service'}]}
        requests.post(f'{HUBSPOT_BASE}/crm/v3/associations/deals/services/batch/create', json=assoc_payload, headers=HUBSPOT_HEADERS())

        contact_ids = get_deal_contact_ids(deal_id)
        associate_service_contacts(service_id, contact_ids)
        create_service_task(service_id, task_details, contact_ids=contact_ids)

        return jsonify({'success': True, 'service_id': service_id, 'name': data.get('name', '')})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/hubspot/services/<service_id>', methods=['PATCH'])
def hs_update_service(service_id):
    try:
        data = request.get_json()
        task_details = data.get('task_details') or data.get('description', '')
        payload = {'properties': build_service_properties(data, include_pipeline=False)}
        if not payload['properties'] and not task_details:
            return jsonify({'error': 'No updatable service properties provided'}), 400

        if payload['properties']:
            res = requests.patch(f'{HUBSPOT_BASE}/crm/v3/objects/services/{service_id}', json=payload, headers=HUBSPOT_HEADERS())
            if res.status_code not in [200, 204]:
                return jsonify({'error': res.text}), res.status_code

        if task_details:
            create_service_task(service_id, task_details)

        return jsonify({'success': True, 'service_id': service_id})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ── Customer-facing pre-installation farm survey ────────────
def _note_association_payload(note_id, obj_id, to_object_type, association_type):
    return {
        'inputs': [{
            'from': {'id': note_id},
            'to': {'id': obj_id},
            'type': association_type
        }]
    }


def _build_checklist_note(data):
    operation_type = data.get('operation_type', '')
    if operation_type == 'Other' and data.get('op_other_text'):
        operation_type = f"Other: {data['op_other_text']}"

    tensioner_labels = {
        'not_applicable': 'Not applicable — all tensioners are in good condition.',
        'customer_supply': 'Customer will supply and replace tensioners prior to installation.',
        'psd_supply': 'Customer requests PSD to supply replacement tensioners at additional cost.'
    }

    lines = [
        f"<strong>PSD Pre-Installation Farm Survey</strong>",
        f"Power Company: {data.get('power_company', '')}",
        f"County: {data.get('county', '')} | Year Farm Built: {data.get('year_built', '')}",
        f"Operation Type: {operation_type or 'Not specified'}",
        f"Total Barns: {data.get('total_barns', '')} | Total Fans: {data.get('total_fans', '')}",
        "",
        "<strong>Fan Configurations:</strong>"
    ]
    for row in data.get('fan_rows', []):
        lines.append(f"- {row.get('num_fans', '')} fans | {row.get('brand', '')} | {row.get('size', '')}in | {row.get('motor_hp', '')}HP | Pulley: {row.get('pulley_size', '')}")

    lines += [
        "",
        "<strong>Equipment &amp; Condition Checklist:</strong>",
        f"Fan blades in good condition: {data.get('fan_blades_condition', '')}",
        f"Fans securely mounted: {data.get('fans_mounted_secure', '')}",
        f"Shutters open/close freely: {data.get('shutters_operate', '')}",
        f"Hardware free of heavy rust: {data.get('hardware_no_rust', '')}",
        f"Hardware removable without grinding: {data.get('hardware_removable', '')}",
        f"Belts inspected: {data.get('belts_inspected', '')}",
        f"Spare belts on hand: {data.get('spare_belts_onhand', '')}",
        f"Tensioners in good condition: {data.get('tensioners_good', '')}",
        f"Bearings in good condition: {data.get('bearings_good', '')}",
        f"Tensioner replacement option: {tensioner_labels.get(data.get('tensioner_option', ''), data.get('tensioner_option', ''))}",
        f"Motors mounted upright: {data.get('motors_upright', '')}",
        f"Motor hardware accessible: {data.get('motor_hardware_accessible', '')}",
        "",
        f"Overall readiness confirmed: {data.get('readiness_confirmed', '') or 'No'}",
        "",
        "<strong>Installation Scheduling:</strong>",
        f"Available Window: {data.get('install_window_from', '')} to {data.get('install_window_to', '')}",
        f"Flock Cycle (weeks between flocks): {data.get('flock_cycle_weeks', '')}",
        f"Scheduling Restrictions: {data.get('scheduling_restrictions', '')}"
    ]
    return '<br>'.join(lines)


def _date_to_hs_ms(date_str):
    if not date_str:
        return None
    try:
        dt = datetime.strptime(date_str, '%Y-%m-%d').replace(tzinfo=timezone.utc)
        return str(int(dt.timestamp() * 1000))
    except ValueError:
        return None


@app.route('/submit-customer-form', methods=['POST'])
def submit_customer_form():
    try:
        data = request.get_json()
        if not data:
            return jsonify({'error': 'No JSON data provided'}), 400
        if CUSTOMER_INTAKE_PIN and str(data.get('access_pin', '')) != CUSTOMER_INTAKE_PIN:
            return jsonify({'error': 'Invalid or missing access PIN'}), 401
        required = ['gp_account_name', 'gp_account_number', 'first_name', 'last_name', 'email', 'farm_address', 'power_company']
        missing = [f for f in required if not data.get(f)]
        if missing:
            return jsonify({'error': f'Missing required fields: {missing}'}), 400

        first_name = data.get('first_name', '').strip()
        last_name = data.get('last_name', '').strip()
        farm_state = POWER_COMPANY_STATE_MAP.get(data.get('power_company', ''), data.get('state', ''))
        operation_type = data.get('operation_type', '')
        if operation_type == 'Other' and data.get('op_other_text'):
            operation_type = data['op_other_text']

        contact_payload = {
            'properties': {
                'email': data.get('email', ''),
                'firstname': first_name,
                'lastname': last_name,
                'phone': data.get('phone', ''),
                'company': data.get('gp_account_name', ''),
                'address': data.get('farm_address', ''),
                'city': data.get('city', ''),
                'state': farm_state,
                'zip': data.get('zip', ''),
                'gp_account_number': str(data.get('gp_account_number', '')),
                'farm_name': data.get('gp_account_name', ''),
                'farm_county': data.get('county', ''),
                'year_farm_built': str(data.get('year_built', '')),
                'total_number_of_barns': str(data.get('total_barns', '')),
                'total_number_of_fans': str(data.get('total_fans', '')),
            }
        }
        contact_payload['properties']['invitation'] = CUSTOMER_INTAKE_INVITATION
        contact_search = {
            'filterGroups': [{'filters': [{'propertyName': 'email', 'operator': 'EQ', 'value': data.get('email', '')}]}],
            'properties': ['email'],
            'limit': 1
        }
        search_res = requests.post(f'{HUBSPOT_BASE}/crm/v3/objects/contacts/search', json=contact_search, headers=HUBSPOT_HEADERS())
        existing_contact = search_res.json().get('results', []) if search_res.status_code == 200 else []
        if existing_contact:
            contact_id = existing_contact[0]['id']
            contact_res = requests.patch(f'{HUBSPOT_BASE}/crm/v3/objects/contacts/{contact_id}', json=contact_payload, headers=HUBSPOT_HEADERS())
            if contact_res.status_code not in [200, 204]:
                return jsonify({'error': contact_res.text}), contact_res.status_code
        else:
            contact_res = requests.post(f'{HUBSPOT_BASE}/crm/v3/objects/contacts', json=contact_payload, headers=HUBSPOT_HEADERS())
            if contact_res.status_code != 201:
                return jsonify({'error': contact_res.text}), contact_res.status_code
            contact_id = contact_res.json()['id']

        deal_payload = {
            'properties': {
                'dealname': f"{data.get('gp_account_name', '')} — {data.get('city', '')}",
                'dealstage': CUSTOMER_FORM_SENT_STAGE,
                'pipeline': CUSTOMER_INTAKE_PIPELINE,
                'total_number_of_barns': str(data.get('total_barns', '')),
                'total_number_of_fans': str(data.get('total_fans', '')),
                'install_window_from': data.get('install_window_from', ''),
                'install_window_to': data.get('install_window_to', ''),
                'flock_cycle_weeks': str(data.get('flock_cycle_weeks', '')),
                'scheduling_restrictions': data.get('scheduling_restrictions', ''),
                'farm_name': data.get('gp_account_name', ''),
                'farm_county': data.get('county', ''),
                'farm_address': data.get('farm_address', ''),
                'state': farm_state,
                'farm_zip': data.get('zip', ''),
                'first_name': first_name,
                'last_name': last_name,
                'phone_number': data.get('phone', ''),
                'customer_email': data.get('email', ''),
                'year_farm_built': str(data.get('year_built', '')),
                'type_of_poultry_opperation': operation_type,
                'hubspot_owner_id': OWNER_ID,
            }
        }
        deal_res = requests.post(f'{HUBSPOT_BASE}/crm/v3/objects/deals', json=deal_payload, headers=HUBSPOT_HEADERS())
        if deal_res.status_code != 201:
            return jsonify({'error': deal_res.text}), deal_res.status_code
        deal_id = deal_res.json()['id']

        assoc_payload = {'inputs': [{'from': {'id': deal_id}, 'to': {'id': contact_id}, 'type': 'deal_to_contact'}]}
        requests.post(f'{HUBSPOT_BASE}/crm/v3/associations/deals/contacts/batch/create', json=assoc_payload, headers=HUBSPOT_HEADERS())

        # Appointment for the installation window (best-effort, does not block submission)
        try:
            appointment_payload = {
                'properties': {
                    'hs_appointment_name': f"{data.get('gp_account_name', '')} Scheduled",
                    'hs_appointment_start': _date_to_hs_ms(data.get('install_window_from', '')),
                    'hs_appointment_end': _date_to_hs_ms(data.get('install_window_to', '')),
                    'hs_pipeline_stage': '83b59094-da00-4f58-bc51-ba4d9a66a248',
                }
            }
            appt_res = requests.post(f'{HUBSPOT_BASE}/crm/v3/objects/appointments', json=appointment_payload, headers=HUBSPOT_HEADERS())
            if appt_res.status_code == 201:
                appointment_id = appt_res.json()['id']
                requests.put(f'{HUBSPOT_BASE}/crm/v4/objects/appointments/{appointment_id}/associations/default/deals/{deal_id}', headers=HUBSPOT_HEADERS())
        except Exception:
            pass

        # Fan configuration rows → line items (best-effort, does not block submission)
        total_fans_qty = 0
        total_line_amount = 0.0
        for row in data.get('fan_rows', []):
            if not (row.get('brand') or row.get('size')):
                continue
            try:
                qty = int(row.get('num_fans') or 1)
            except (TypeError, ValueError):
                qty = 1
            total_fans_qty += qty
            try:
                name_parts = [row.get('brand', ''), row.get('size', ''), row.get('motor_hp', ''), row.get('pulley_size', '')]
                li_payload = {
                    'properties': {
                        'name': ' '.join(p for p in name_parts if p),
                        'quantity': str(qty),
                        'price': UNIT_PRICE,
                        'fan_brand': row.get('brand', ''),
                        'fan_size': str(row.get('size', '')),
                        'motor_size': str(row.get('motor_hp', '')),
                        'description': f"Pulley size: {row.get('pulley_size', '')}",
                    }
                }
                li_res = requests.post(f'{HUBSPOT_BASE}/crm/v3/objects/line_items', json=li_payload, headers=HUBSPOT_HEADERS())
                if li_res.status_code == 201:
                    li_id = li_res.json()['id']
                    requests.post(
                        f'{HUBSPOT_BASE}/crm/v3/associations/deals/line_items/batch/create',
                        json={'inputs': [{'from': {'id': deal_id}, 'to': {'id': li_id}, 'type': 'deal_to_line_item'}]},
                        headers=HUBSPOT_HEADERS()
                    )
                    total_line_amount += qty * float(UNIT_PRICE)
            except Exception:
                pass

        # Installation charges line item, quantity = total fans across all lines (best-effort)
        try:
            install_qty = total_fans_qty or 1
            install_unit_price = 4
            install_payload = {
                'properties': {
                    'name': 'Installation Charges',
                    'quantity': str(install_qty),
                    'price': str(install_unit_price),
                }
            }
            install_res = requests.post(f'{HUBSPOT_BASE}/crm/v3/objects/line_items', json=install_payload, headers=HUBSPOT_HEADERS())
            if install_res.status_code == 201:
                install_id = install_res.json()['id']
                requests.post(
                    f'{HUBSPOT_BASE}/crm/v3/associations/deals/line_items/batch/create',
                    json={'inputs': [{'from': {'id': deal_id}, 'to': {'id': install_id}, 'type': 'deal_to_line_item'}]},
                    headers=HUBSPOT_HEADERS()
                )
                total_line_amount += install_qty * install_unit_price
        except Exception:
            pass

        # Update deal amount to reflect all line items (best-effort)
        try:
            requests.patch(
                f'{HUBSPOT_BASE}/crm/v3/objects/deals/{deal_id}',
                json={'properties': {'amount': str(total_line_amount)}},
                headers=HUBSPOT_HEADERS()
            )
        except Exception:
            pass

        # Checklist + survey details → note on the deal and contact (best-effort)
        try:
            note_payload = {'properties': {'hs_note_body': _build_checklist_note(data), 'hs_timestamp': str(int(time.time() * 1000))}}
            note_res = requests.post(f'{HUBSPOT_BASE}/crm/v3/objects/notes', json=note_payload, headers=HUBSPOT_HEADERS())
            if note_res.status_code == 201:
                note_id = note_res.json()['id']
                requests.put(f'{HUBSPOT_BASE}/crm/v4/objects/notes/{note_id}/associations/default/deals/{deal_id}', headers=HUBSPOT_HEADERS())
                requests.put(f'{HUBSPOT_BASE}/crm/v4/objects/notes/{note_id}/associations/default/contacts/{contact_id}', headers=HUBSPOT_HEADERS())
        except Exception:
            pass

        return jsonify({'success': True, 'contact_id': contact_id, 'deal_id': deal_id})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


def _safe_folder_segment(value):
    cleaned = re.sub(r'[^A-Za-z0-9 _-]', '', str(value or '')).strip().replace(' ', '_')
    return cleaned[:60]


@app.errorhandler(413)
def payload_too_large(_e):
    return jsonify({'error': 'Upload is too large.'}), 413


@app.route('/customer-intake-upload', methods=['POST'])
def customer_intake_upload():
    """Upload optional intake documents to HubSpot Files and attach them to the deal.

    Called after /submit-customer-form has created the deal. Files land in the
    HubSpot file manager as PRIVATE, then a single note carrying every file id
    in hs_attachment_ids is associated to the deal.
    """
    try:
        if CUSTOMER_INTAKE_PIN and str(request.form.get('access_pin', '')) != CUSTOMER_INTAKE_PIN:
            return jsonify({'error': 'Invalid or missing access PIN'}), 401

        deal_id = str(request.form.get('deal_id', '')).strip()
        if not deal_id.isdigit():
            return jsonify({'error': 'A valid deal_id is required'}), 400

        uploads = [f for f in request.files.getlist('files') if f and f.filename]
        if not uploads:
            return jsonify({'error': 'No files provided'}), 400
        if len(uploads) > MAX_UPLOAD_FILES:
            return jsonify({'error': f'At most {MAX_UPLOAD_FILES} files may be uploaded at once'}), 400

        folder_name = _safe_folder_segment(request.form.get('account_name', '')) or deal_id
        folder_path = f'{INTAKE_FILES_FOLDER}/{folder_name}'
        # Content-Type is deliberately omitted so requests sets the multipart boundary.
        auth_header = {'Authorization': f'Bearer {HUBSPOT_API_KEY}'}

        uploaded = []
        failed = []
        for upload in uploads:
            filename = os.path.basename(upload.filename)
            blob = upload.read()
            if not blob or len(blob) > MAX_UPLOAD_BYTES:
                failed.append(filename)
                continue
            try:
                res = requests.post(
                    HUBSPOT_FILES_URL,
                    headers=auth_header,
                    files={'file': (filename, io.BytesIO(blob), upload.mimetype or 'application/octet-stream')},
                    data={
                        'options': json.dumps({'access': 'PRIVATE'}),
                        'folderPath': folder_path,
                    },
                    timeout=120
                )
                file_id = str(res.json().get('id', '')) if res.status_code in (200, 201) else ''
                if file_id:
                    uploaded.append({'id': file_id, 'name': filename})
                else:
                    failed.append(filename)
            except Exception:
                failed.append(filename)

        note_id = None
        if uploaded:
            note_body = '<br>'.join(
                ['<strong>Customer-uploaded documents</strong>'] + [f"- {f['name']}" for f in uploaded]
            )
            note_payload = {
                'associations': [{
                    'to': {'id': deal_id},
                    'types': [{
                        'associationCategory': 'HUBSPOT_DEFINED',
                        'associationTypeId': NOTE_TO_DEAL_ASSOCIATION_TYPE_ID
                    }]
                }],
                'properties': {
                    'hs_note_body': note_body,
                    'hs_timestamp': datetime.now(timezone.utc).strftime('%Y-%m-%dT%H:%M:%SZ'),
                    'hs_attachment_ids': ';'.join(f['id'] for f in uploaded),
                }
            }
            note_res = requests.post(f'{HUBSPOT_BASE}/crm/v3/objects/notes', json=note_payload, headers=HUBSPOT_HEADERS())
            if note_res.status_code != 201:
                return jsonify({'error': f'Files uploaded but could not be attached to the deal: {note_res.text}'}), 502
            note_id = note_res.json().get('id')

        return jsonify({
            'success': True,
            'deal_id': deal_id,
            'note_id': note_id,
            'uploaded': [f['name'] for f in uploaded],
            'failed': failed,
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ── Proxy to Make webhooks (intake pipeline only) ───────────
@app.route('/proxy', methods=['POST'])
def proxy():
    try:
        data = request.get_json()
        target_url = data.get('url')
        payload = data.get('payload')

        if not target_url:
            return jsonify({'error': 'Missing target url'}), 400

        response = requests.post(target_url, json=payload, headers={'Content-Type': 'application/json'}, timeout=30)

        content_type = response.headers.get('Content-Type', '')
        if 'text/html' in content_type or response.text.strip().startswith('<'):
            return jsonify({'error': 'Make returned an error', 'status': response.status_code}), 500

        try:
            return jsonify(response.json()), response.status_code
        except Exception:
            try:
                cleaned = response.text.strip()
                cleaned = re.sub(r',\s*}', '}', cleaned)
                cleaned = re.sub(r',\s*]', ']', cleaned)
                return jsonify(json.loads(cleaned)), response.status_code
            except Exception:
                return jsonify({'raw': response.text}), response.status_code

    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ── Proxy to Anthropic API ───────────────────────────────────
@app.route('/claude', methods=['POST'])
def claude_proxy():
    try:
        data = request.get_json()
        response = requests.post(
            'https://api.anthropic.com/v1/messages',
            json=data,
            headers={'Content-Type': 'application/json', 'x-api-key': ANTHROPIC_API_KEY, 'anthropic-version': '2023-06-01'},
            timeout=60
        )
        return jsonify(response.json()), response.status_code
    except Exception as e:
        return jsonify({'error': str(e)}), 500

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=False)
